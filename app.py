from flask import Flask, render_template, request, redirect, url_for, flash, send_file, session
import pandas as pd
import os
import plotly

app = Flask(__name__)
app.secret_key = 'your_secret_key'
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def compute_changes(df_actual, df_cleaned):
    """
    Compare the original data (df_actual) with the cleaned data (df_cleaned)
    and return a dictionary summarizing the changes.
    """
    changes = {}
    duplicate_count = int(df_actual.duplicated().sum())
    changes["duplicates_removed"] = duplicate_count

    missing_before = df_actual.isnull().sum()
    missing_after = df_cleaned.isnull().sum()
    missing_filled = (missing_before - missing_after).to_dict()
    changes["missing_values_filled"] = missing_filled

    return changes

def header_detected(df):
    """
    Check if the DataFrame's columns indicate that headers were properly detected.
    If every column name starts with 'Unnamed', we assume headers are missing.
    """
    return not all(str(col).startswith("Unnamed") for col in df.columns)

def read_file(file_path):
    """
    Utility function to read CSV or Excel file based on extension.
    If the user has not confirmed to continue without proper headers,
    we check that headers are detected.
    If user confirmed (session flag 'header_confirm' is True), we force the first row as header.
    """
    if file_path.lower().endswith('.csv'):
        # If confirmed, re-read with header=0; header=0 is default but we specify it explicitly.
        df = pd.read_csv(file_path, header=0) if session.get('header_confirm', False) else pd.read_csv(file_path)
    else:
        df = pd.read_excel(file_path, header=0) if session.get('header_confirm', False) else pd.read_excel(file_path)
        
    if not session.get('header_confirm', False):
        if not header_detected(df):
            raise Exception("Header row not detected. Please confirm if you want to treat the first row as header.")
    return df

def read_excel_sheet(file_path, expected_sheet):
    """
    Try to read the expected sheet from an Excel file.
    If not found, fallback to the first sheet.
    Also verifies that the header row exists (unless the user confirmed to continue).
    """
    try:
        xls = pd.ExcelFile(file_path)
        if expected_sheet in xls.sheet_names:
            df = pd.read_excel(file_path, sheet_name=expected_sheet)
        else:
            df = pd.read_excel(file_path, sheet_name=xls.sheet_names[0])
        if not session.get('header_confirm', False):
            if not header_detected(df):
                raise Exception("Header row not detected in Excel sheet. Please confirm if you want to treat the first row as header.")
        return df
    except Exception as e:
        raise Exception(f"Error reading Excel file: {e}")

def remove_extra_sheets(file_path, allowed_sheets=["Actual Data", "Cleaned Data", "Reports"]):
    """
    Remove any sheets in the Excel file that are not in the allowed list.
    """
    from openpyxl import load_workbook
    wb = load_workbook(file_path)
    for sheet in wb.sheetnames:
        if sheet not in allowed_sheets:
            wb.remove(wb[sheet])
    wb.save(file_path)

@app.route('/', methods=['GET', 'POST'])
def index():
    # Clear any previous header confirmation.
    session.pop('header_confirm', None)
    if request.method == 'POST':
        if 'data_file' not in request.files:
            flash("Please upload a file.")
            return redirect(request.url)
        file = request.files['data_file']
        if file.filename == '':
            flash("Please upload a file.")
            return redirect(request.url)
        
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)
        
        try:
            df = read_file(file_path)
            header_missing = False
        except Exception as e:
            header_missing = True
            flash(str(e))
            # Render the page with a confirmation prompt.
            return render_template('index.html', file_path=file_path, file_name=file.filename, header_missing=True)
        
        missing_values = df.isnull().sum().to_dict()
        duplicate_rows = int(df.duplicated().sum())
        cleaning_summary = {
            'missing_values': missing_values,
            'duplicate_rows': duplicate_rows
        }
        
        return render_template('index.html', 
                               file_path=file_path, 
                               file_name=file.filename,
                               cleaning_summary=cleaning_summary, 
                               df_html=df.head().to_html())
    return render_template('index.html')

@app.route('/confirm_header', methods=['POST'])
def confirm_header():
    """
    If the user confirms that the first row should be treated as header,
    set a session flag and re-read the file with header=0.
    """
    file_path = request.form.get('file_path')
    session['header_confirm'] = True
    try:
        if file_path.lower().endswith('.csv'):
            df = pd.read_csv(file_path, header=0)
        else:
            df = pd.read_excel(file_path, header=0)
    except Exception as e:
        flash(f"Error reading file after confirmation: {e}")
        return redirect(url_for('index'))
    
    missing_values = df.isnull().sum().to_dict()
    duplicate_rows = int(df.duplicated().sum())
    cleaning_summary = {
        'missing_values': missing_values,
        'duplicate_rows': duplicate_rows
    }
    flash("Continuing with first row as header.")
    return render_template('index.html', 
                           file_path=file_path, 
                           file_name=os.path.basename(file_path),
                           cleaning_summary=cleaning_summary, 
                           df_html=df.head().to_html(),
                           header_confirmed=True)

@app.route('/clean', methods=['POST'])
def clean_file():
    file_path = request.form.get('file_path')
    if not file_path or not os.path.exists(file_path):
        flash("No file uploaded. Please upload file.")
        return redirect(url_for('index'))
    
    try:
        df = read_file(file_path)
    except Exception as e:
        flash(f"Error reading file: {e}")
        return redirect(url_for('index'))
    
    df_cleaned = df.drop_duplicates()
    df_cleaned.fillna(method='ffill', inplace=True)
    
    try:
        if file_path.lower().endswith('.csv'):
            actual_path = file_path.replace('.csv', '_actual.csv')
            cleaned_path = file_path.replace('.csv', '_cleaned.csv')
            df.to_csv(actual_path, index=False)
            df_cleaned.to_csv(cleaned_path, index=False)
            file_path = cleaned_path
        else:
            writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
            df.to_excel(writer, sheet_name='Actual Data', index=False)
            df_cleaned.to_excel(writer, sheet_name='Cleaned Data', index=False)
            writer.close()
            remove_extra_sheets(file_path, allowed_sheets=["Actual Data", "Cleaned Data"])
    except Exception as e:
        flash(f"Error during cleaning process: {e}")
        return redirect(url_for('index'))
    
    flash("Cleaning completed. You can now generate reports.")
    return render_template('index.html', 
                           file_path=file_path, 
                           file_name=os.path.basename(file_path),
                           cleaned_data_html=df_cleaned.head().to_html())

import matplotlib.pyplot as plt

REPORT_FOLDER = "reports"
IMAGE_FOLDER = "static"
os.makedirs(REPORT_FOLDER, exist_ok=True)

def process_excel(file_path):
    """
    Process the Excel file and generate a report with three sheets:
    Actual Data, Cleaned Data, and Reports.
    """
    df = read_file(file_path)
    df_cleaned = df.drop_duplicates().dropna()
    report_path = os.path.join(REPORT_FOLDER, "processed_report.xlsx")
    try:
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name="Actual Data", index=False)
            df_cleaned.to_excel(writer, sheet_name="Cleaned Data", index=False)
            report_df = pd.DataFrame({
                "Total Rows (Original)": [df.shape[0]],
                "Total Rows (Cleaned)": [df_cleaned.shape[0]],
                "Duplicate Rows Removed": [df.shape[0] - df_cleaned.shape[0]]
            })
            report_df.to_excel(writer, sheet_name="Reports", index=False)
    except Exception as e:
        print(f"Error saving Excel report: {e}")
        return None, None

    image_path = os.path.join(IMAGE_FOLDER, "histogram.png")
    try:
        plt.figure(figsize=(8, 6))
        df_cleaned.hist(bins=20)
        plt.savefig(image_path)
        plt.close()
    except Exception as e:
        print(f"Error generating histogram: {e}")
        return report_path, None

    remove_extra_sheets(report_path)
    return report_path, image_path

@app.route('/report', methods=['POST'])
def generate_report():
    file_path = request.form.get('file_path')
    if not file_path or not os.path.exists(file_path):
        flash("No file uploaded. Please upload file.")
        return redirect(url_for('index'))
    
    try:
        if file_path.lower().endswith('.csv'):
            df_cleaned = pd.read_csv(file_path)
        else:
            df_cleaned = read_excel_sheet(file_path, 'Cleaned Data')
    except Exception as e:
        flash(f"Error reading cleaned data: {e}")
        return redirect(url_for('index'))
    
    report_df = df_cleaned.describe().transpose()
    
    try:
        if file_path.lower().endswith('.csv'):
            report_path = file_path.replace('.csv', '_report.xlsx')
            actual_file = file_path.replace('_cleaned.csv', '_actual.csv')
            try:
                df_actual = pd.read_csv(actual_file)
            except Exception:
                df_actual = pd.read_csv(file_path)
            writer = pd.ExcelWriter(report_path, engine='openpyxl')
            df_actual.to_excel(writer, sheet_name='Actual Data', index=False)
            df_cleaned.to_excel(writer, sheet_name='Cleaned Data', index=False)
            report_df.to_excel(writer, sheet_name='Reports', index=True)
            writer.close()
            file_path = report_path
        else:
            writer = pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace')
            report_df.to_excel(writer, sheet_name='Reports', index=True)
            writer.close()
            remove_extra_sheets(file_path)
    except Exception as e:
        flash(f"Error generating report: {e}")
        return redirect(url_for('index'))
    
    import plotly.express as px
    from openpyxl.drawing.image import Image as OpenpyxlImage
    import tempfile

    if not file_path.lower().endswith('.csv'):
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        reports_sheet = wb['Reports'] if 'Reports' in wb.sheetnames else wb.worksheets[0]
        start_row = report_df.shape[0] + 3
        current_row = start_row
        temp_dir = tempfile.mkdtemp()
        image_files = []
        numeric_cols = df_cleaned.select_dtypes(include=['number']).columns

        for col in numeric_cols:
            fig = px.histogram(df_cleaned, x=col, title=f'Histogram of {col}')
            img_path = os.path.join(temp_dir, f"hist_{col}.png")
            try:
                fig.write_image(img_path)
                image_files.append((img_path, f'Histogram of {col}'))
            except Exception as e:
                flash(f"Error generating histogram for {col}: {e}")
        
        for col in numeric_cols:
            fig = px.box(df_cleaned, y=col, title=f'Box Plot of {col}')
            img_path = os.path.join(temp_dir, f"box_{col}.png")
            try:
                fig.write_image(img_path)
                image_files.append((img_path, f'Box Plot of {col}'))
            except Exception as e:
                flash(f"Error generating box plot for {col}: {e}")
        
        if len(numeric_cols) >= 2:
            fig = px.scatter(df_cleaned, x=numeric_cols[0], y=numeric_cols[1],
                             title=f'Scatter Plot: {numeric_cols[0]} vs {numeric_cols[1]}')
            img_path = os.path.join(temp_dir, "scatter.png")
            try:
                fig.write_image(img_path)
                image_files.append((img_path, f'Scatter Plot: {numeric_cols[0]} vs {numeric_cols[1]}'))
            except Exception as e:
                flash(f"Error generating scatter plot: {e}")
        
        if len(numeric_cols) > 1:
            corr = df_cleaned[numeric_cols].corr()
            fig = px.imshow(corr, text_auto=True, title='Correlation Heatmap')
            img_path = os.path.join(temp_dir, "heatmap.png")
            try:
                fig.write_image(img_path)
                image_files.append((img_path, "Correlation Heatmap"))
            except Exception as e:
                flash(f"Error generating correlation heatmap: {e}")
        
        for img_path, title in image_files:
            reports_sheet.cell(row=current_row, column=1, value=title)
            current_row += 1
            img = OpenpyxlImage(img_path)
            img.anchor = f'A{current_row}'
            reports_sheet.add_image(img)
            current_row += 20

        wb.save(file_path)
    
    flash("Report generated and saved in the file with visuals.")
    return redirect(url_for('download_page', file_path=file_path))

@app.route('/download_page', methods=['GET'])
def download_page():
    file_path = request.args.get('file_path')
    if not file_path or not os.path.exists(file_path):
        flash("File not found.")
        return redirect(url_for('index'))
    
    try:
        if file_path.lower().endswith('.csv'):
            df_actual = pd.read_excel(file_path, sheet_name='Actual Data')
            df_cleaned = pd.read_excel(file_path, sheet_name='Cleaned Data')
        else:
            df_actual = read_excel_sheet(file_path, 'Actual Data')
            df_cleaned = read_excel_sheet(file_path, 'Cleaned Data')
    except Exception as e:
        flash(f"Error reading file sheets: {e}")
        return redirect(url_for('index'))
    
    changes = compute_changes(df_actual, df_cleaned)
    return render_template('download.html', file_path=file_path, changes=changes)

@app.route('/download_file', methods=['GET'])
def download_file():
    file_path = request.args.get('file_path')
    if not file_path or not os.path.exists(file_path):
        flash("File not found.")
        return redirect(url_for('index'))
    return send_file(file_path, as_attachment=True)

@app.route('/clear_flashes', methods=['POST'])
def clear_flashes():
    session.pop('_flashes', None)
    return '', 204

if __name__ == '__main__':
    app.run(debug=True)
